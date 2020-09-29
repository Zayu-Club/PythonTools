import openpyxl
import re
import ast
from urllib.parse import urlparse
import socket

#Only xlsx format is supported
WORKPATH = 'D:\提取IP测试工作簿.xlsx'
SHEETNAME = '咕咕咕'
URLCOL = 2
DSTCOL = 3

#由URL获取域名
def getDomainFromURL(url):
    #若url没有HTTP前缀，则加上前缀
    if(not url.startswith('http')): 
        url = 'http://' + url
    parsed_uri = urlparse(url)
    domain = '{uri.netloc}'.format(uri=parsed_uri)
    #若url有端口号，则从{uri.netloc}中去掉端口号，得到最终域名
    index_port = domain.find(':')
    if(not index_port == -1):
        domain = domain[:-(len(domain) - index_port)]
    print("[INFO]提取域名 " + domain)
    return domain

#由域名获取IP列表
def getIpListFromDomain(domain):
    ip_list = []
    try:
        addr_list = socket.getaddrinfo(domain,None)
        for item in addr_list:
            if(item[4][0] not in ip_list):
                print("[INFO]提取域名 " + domain + " 的IP：" + item[4][0])
                ip_list.append(item[4][0])
        return ip_list
    except Exception as e:
        print("[ERROR]域名解析失败")
        return ['域名解析失败']

#提取IP地址
def getUrl2ip(urlStr):
    #首先尝试直接正则匹配提取IP
    resultIP = re.findall(r'(?:(?:25[0-5]|2[0-4]\d|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)', urlStr)
    if(resultIP):
        #print("[INFO]提取结果：" + resultIP)
        return resultIP
    else:
        #以域名方式提取IP
        print("[WARNING]提取IP失败，尝试以域名提取")
        resultIPList = getIpListFromDomain(getDomainFromURL(urlStr))
        return resultIPList

def main(): 
    try:
        #打开工作簿
        workBook = openpyxl.load_workbook(WORKPATH,data_only=True)    
        print("[INFO]工作簿打开成功：" + WORKPATH + "," + str(workBook.sheetnames))
    except Exception as e:
        print("[ERROR]工作簿打开失败：" + WORKPATH + "," + str(workBook.sheetnames))
        return
        #选择工作表
    try:
        workSheet = workBook[SHEETNAME]
        print("[INFO]目标工作表选择：" + SHEETNAME)
    except Exception as e:
        print("[ERROR]工作表打开失败：")
        return    
    #获取最大列&最大行
    maxCol = workSheet.max_column
    maxRow = workSheet.max_row
    print("[INFO]获取工作表最大列、行：" + str(maxCol) + "，" + str(maxRow))
    #遍历提取IP写入dstCol列中
    for i in range(2,maxRow+1):
        urlStr = workSheet.cell(row=i, column=URLCOL).value
        if urlStr:
            print("[INFO]读取的URL：" + str(urlStr))
            resList = getUrl2ip(str(urlStr))
            directionCell = workSheet.cell(row=i, column=DSTCOL)
            res = ''
            for item in resList:
                res += str(item) + '\n'
            directionCell.value = res
            print("[INFO]写入IP:")
    workBook.save(WORKPATH)

if __name__ == '__main__':
    args = 'n'
    print("[INFO]目标工作簿：" + WORKPATH)
    print("[INFO]目标工作表：" + SHEETNAME)
    print("[INFO]URL列号：" + str(URLCOL))
    print("[INFO]IP列号：" + str(DSTCOL))
    while not args == 'y' and not args == 'Y':
        args = input("[WARNING]对工作簿的操作无法恢复，请在操作前备份工作簿，按Y键继续[N]：")
    main()
    print("[INFO]全部完成")